from openpyxl import load_workbook
import matplotlib.pyplot as plt
import math


def calc_score(company, time):
	score = company['BP'][time].value + company['EP'][time].value + company['ROE'][time].value / 100.0 + \
			company['PROFIT_GROWTH_YOY'][time].value - company['VOL120'][time].value / 100.0
	if abs(company['PRICE'][time].value < 1e-5):
		score = -100
	return score

def get_score(comp):
	return comp['score']

def calc_ratio(company, time):
	value_sum = 0
	for comp in company:
		if company[comp]['bought_time'] > 0:
			value_sum += company[comp]['VALUE'][time].value
	for comp in company:
		if company[comp]['bought_time'] > 0:
			company[comp]['ratio'] = company[comp]['VALUE'][time].value * 1.0 / value_sum
			if company[comp]['ratio'] > 0.05:
				company[comp]['ratio'] = 0.05
	return

def single_rate(company, time):
	for comp in company:
		if company[comp]['bought_time'] > 0:
			if company[comp]['bought_time'] == time:
				company[comp]['rate'] = 0
				continue
			delta_t = time - company[comp]['bought_time']
			bought_price = company[comp]['PRICE'][company[comp]['bought_time']].value
			present_price = company[comp]['PRICE'][time].value
			company[comp]['rate'] = math.pow((present_price * 1.0 / bought_price), 1.0 / delta_t) - 1
	return

def total_rate(company, time):
	rate_sum = 0
	for comp in company:
		if company[comp]['bought_time'] > 0:
			rate_sum += company[comp]['rate'] * company[comp]['ratio']
	return rate_sum

def top_k(company, time, k=200):
	company_list = []
	for comp in company:
		company_list.append(company[comp])
	company_list.sort(key=get_score, reverse=True)
	thres = company_list[k - 1]['score']
	for comp in company:
		if company[comp]['score'] > thres:
			if company[comp]['bought_time'] > 0:
				continue
			else:
				company[comp]['bought_time'] = time
		else:
			company[comp]['bought_time'] = 0
	return


if __name__ == '__main__':
	
	# read xlsx into a dictionary
	data = load_workbook('data.xlsx')
	sheet = data.get_sheet_by_name('Sheet1')
	company = {}
	line = 0
	company_name = ''
	for row in sheet.rows:
		line += 1
		''' read original kind of data
		if line % 7 == 3:
			company_name = row[0].value
			company[company_name] = {}
			company[company_name]['BP'] = row[2:23]
			company[company_name]['score'] = 0
			company[company_name]['bought_time'] = 0
			company[company_name]['ratio'] = 0
			company[company_name]['rate'] = 0
			continue

		elif line % 7 == 4:
			company[company_name]['EP'] = row[2:23]
			continue

		elif line % 7 == 5:
			company[company_name]['ROE'] = row[2:23]
			continue

		elif line % 7 == 6:
			company[company_name]['PRICE'] = row[2:23]
			continue

		elif line % 7 == 0:
			company[company_name]['PROFIT_GROWTH_YOY'] = row[2:23]
			continue

		elif line % 7 == 1:
			if line == 1:
				continue
			company[company_name]['VOL120'] = row[2:23]
			continue

		elif line % 7 == 2:
			if line == 2:
				continue
			company[company_name]['VALUE'] = row[2:23]
			continue
		'''
		# read new kind of data
		if line <= 2 or line > 3662:
			continue
		company_code = row[0].value
		company_name = row[1].value
		company[company_name] = {}
		company[company_name]['BP'] = row[2:23]
		company[company_name]['EP'] = row[23:44]
		company[company_name]['ROE'] = row[44:65]
		company[company_name]['PROFIT_GROWTH_YOY'] = row[65:86]
		company[company_name]['VOL120'] = row[86:107]
		company[company_name]['VALUE'] = row[107:128]
		company[company_name]['PRICE'] = row[128:149]
		company[company_name]['score'] = 0
		company[company_name]['bought_time'] = 0
		company[company_name]['ratio'] = 0
		company[company_name]['rate'] = 0

	# from time1 to time21
	ans = []
	for time in range(0, 21):
		for comp in company:
			company[comp]['score'] = calc_score(company[comp], time)
		# print(company['浦发银行']['score'])
		top_k(company, time, 200)

		# skip time1
		if time == 0:
			continue
		calc_ratio(company, time)
		single_rate(company, time)
		ans.append(total_rate(company, time))
		print('time {0}: {1:.8f}'.format(time+1, ans[time-1]))

	# plot
	time = []
	for i in range(2009, 2019):
		date = '{}.6.30'.format(i)
		time.append(date)
		date = '{}.12.31'.format(i)
		time.append(date)
	# print(time)

	hushen = load_workbook('hushen.xlsx', data_only=True)
	hushen_sheet = hushen.get_sheet_by_name('Sheet1')
	hushen_rate = []
	hushen_date = []
	for row in hushen_sheet.rows:
		if row[0].value == '时间':
			continue
		year, month, day = int(row[0].value.split('-')[0]), int(row[0].value.split('-')[1]), int(row[0].value.split('-')[2])
		hushen_date.append('{}.{}.{}'.format(year, month, day))
		hushen_rate.append(float(row[2].value)*100)
	hushen_date.reverse()
	hushen_rate.reverse()
	# print(hushen_date)

	plt.plot(time, ans, 'gs', hushen_date, hushen_rate, 'r^')
	plt.show()
